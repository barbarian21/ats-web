import threading
import json
import re
import datetime
from multiprocessing import Process,Lock
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, colors, Alignment, numbers
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from django.http import HttpResponse, HttpResponseRedirect, FileResponse,JsonResponse
from django.shortcuts import render
from django.views.generic.base import View
from django.conf import settings

from testCoverage.dataDeal import *
from testCoverage.models import TestCoverage
from common.models import Git as Git_Object
from common.models import Station,Project
# -*- coding:utf-8 -*-


#静态变量，所有案子的数据  
#key值为project_git_branch, value为所有站位的数据，路径，测项列表，命令列表
# allproject_Result={} 
#进程的锁，保存数据              
lock=Lock()

#根据测项名或命令查找该测项的所有信息
class SearchView(View):
    def post(self,request):
        allproject_Result=settings.SITE_ALLPROJECTINFO
        context={}
        searchList = None

        inputStr = request.POST['inputStr']
        mark = request.POST['mark']
        project=request.POST['project']
        git=request.POST['git']
        branch=request.POST['branch']

        key_str=project+"_"+git+"_"+branch                      #拼出key值
        thisProject_Result=allproject_Result[key_str]['TC']     #得到这个案子的数据
        lowerInput = inputStr.lower()                           # lower input value
        regex = r'(%s)' % lowerInput                            # create regex
        resultList = []                                         #保存查询到的数据

        if not mark == 'Item':
            searchList = list(set(allproject_Result[key_str]['command']))   # for repeat command
        else:
            searchList = allproject_Result[key_str]['item']

        for item in searchList:
            # 解析的过程中拿到所有的[station:item:command] SearchList
            # sperate by :
            info = item.split(':')
            sta = info[0]
            name = info[1]
            command = None
            lowerName = name.lower()

            if not mark == 'Item':
                command = info[2]
            if not mark == 'Item':
                lowerName = command.lower()

            # catch by regex
            regexRes = re.search(regex, lowerName)

            if regexRes:
                stationInfo = thisProject_Result[sta]           # get item info from all info
                itemInfo = self.getItemInfo(stationInfo, name,sta)   #得到sta站位，测向名为name的所有数据
                resultList.append(itemInfo)         

        # return resultList
        context["resultList"]= resultList
        return render(request, 'testCoverage/search_result.html', context)

    def getItemInfo(self,infoArr, itemName,stationName):
        if not infoArr:
            return {}

        for item in infoArr:
            if item['name'] == itemName:
                item['station'] = stationName
                return item

class TCIndexView(View):
    def get(self,request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect('/login/')
        else:
            return render(request, 'testCoverage/tc_home.html')
            
    def post(self,request):
        if not request.user.is_authenticated:
            return HttpResponseRedirect('/login/')
        else:
            return render(request, 'testCoverage/tc_home.html')


class StationView(View):
    def get(self,request):
        allproject_Result=settings.SITE_ALLPROJECTINFO
        context={}
        station_tc=[]

        project=request.GET.get('project')
        git=request.GET.get('git')
        branch=request.GET.get('branch')
        request_station=request.GET.get('station')
        page = request.GET.get('page')

        # for k in allproject_Result.keys():
        #     print(k)
            
        key_str=project+"_"+git+"_"+branch
        thisProject_Result=allproject_Result[key_str]['TC']
        
        #查询TestCoverage表，将每个测项的description一并返回
        if request_station:
            # project_code = Project.objects.get(code=project)
            # station = Station.objects.get(script=request_station,project=project_code)
            # for i in thisProject_Result[station.script]:
            #     description=TestCoverage.objects.filter(project=project,station=station,test_name=i['name']).values('description')
            #     if description:
            #         i['description']=description[0]['description']
            #     else:
            #         i['description']=''
            station_tc = thisProject_Result[request_station]
            context['station_tc'] = station_tc

            #根据请求的page，返回不同页面
            if page == 'second':
                return render(request, 'testCoverage/tech_csv.html', context)
            else:
                return render(request, 'testCoverage/main_limit_csv.html', context)


class ItemTechView(View):
    def get(self,request):
        allproject_Result=settings.SITE_ALLPROJECTINFO
        context = {}
        item_tech = []

        project=request.GET.get('project')
        git=request.GET.get('git')
        branch=request.GET.get('branch')
        request_station = request.GET.get('station')
        request_item = request.GET.get('item')

        #从文件读取特定分支的数据
        key_str=project+"_"+git+"_"+branch
        thisStation_Result=allproject_Result[key_str]['TC']

        for item in thisStation_Result[request_station]:
            if item['name'] == request_item:
                item_tech = item['logic']
                break

        if item_tech:
            #重新拼接测项名和itemtech
            infoTech = [item_tech[0]]
            for line in item_tech[1:]:
                line['TestName']=''
                infoTech.append(line)
            context['station_tc'] = infoTech
        else:
            context['station_tc'] = []

        return render(request, 'testCoverage/item_tech.html',context)


class ItemLimitView(View):
    def post(self,request):
        allproject_Result=settings.SITE_ALLPROJECTINFO

        project = request.POST['project']
        git=request.POST['git']
        branch=request.POST['branch']
        request_station = request.POST['station']
        request_item = request.POST['item']

        #从文件读取特定分支的数据
        key_str=project+"_"+git+"_"+branch
        thisStation_Result=allproject_Result[key_str]['TC']
        for item in thisStation_Result[request_station]:
            if item['name'] == request_item:
                item_limit = item['limit']
                break
            else:
                item_limit={}  

        return JsonResponse({'itemLimit':item_limit}, safe=True)


class InitTC(View):
    def post(self,request):         #选择project,选择Git，选择branch时，调用
        context={}
        branches=[]
        gits=[]
        stationlist={}
        project = request.user.favor_project
        git=request.POST['git']
        branch=request.POST['branch']
        
        result_dic=self.loadInfo(project,git,branch)
        
        return JsonResponse(result_dic, safe=True)

    def get(self,request):        
        context={}
        project = request.user.favor_project
        git=""
        branch=""

        if not request.user.is_authenticated:
            return HttpResponseRedirect('/login/')
        else:
            context=self.loadInfo(project,git,branch)        
            return render(request, 'testCoverage/tc_home.html',context)

    def loadInfo(self,project,git,branch):
        allproject_Result=settings.SITE_ALLPROJECTINFO
        now_branch=''
        now_git=""
        branches=[]
        gits=[]
        stationlist={}
        result_dic={}

        project=str(project)
        #如果Git为空，加载数据库中的所有Git，并默认为local-git,并得到所有分支
        if git=='':
            allGit=Git_Object.objects.filter(project=project).values('git')
            for i in allGit:
                gits.append(i['git'])
            gits=list(set(gits))                    #去除重复的git
            now_git='local-git' if "local-git" in gits else gits[0]

            allBranch=Git_Object.objects.filter(project=project,git=now_git).values('branch')
            for i in allBranch:
                branches.append(i['branch'])
            now_branch=branches[0]
        else:
            now_git=git
            #如果分支为空的话，分支默认为表中该案子第一个branch,并返回branchs给前端
            if branch=='':
                allBranch=Git_Object.objects.filter(project=project,git=git).values('branch')
                for i in allBranch:
                    branches.append(i['branch'])
                now_branch=branches[0]
            else:
                now_branch=branch

        #当全局变量allproject_Result中没有需要的数据，根据信息查询Git,拿到地址，去解析
        key_str=project+"_"+now_git+"_"+now_branch
        if key_str in allproject_Result.keys():
            thisProject_Result=allproject_Result[key_str]["TC"]
        else:
            parseTCData(project,now_git,now_branch)
            thisProject_Result=allproject_Result[key_str]["TC"]

        #站位根据分类进行查询，查询Station表中的站位，使用字段script匹配
        stations=Station.objects.filter(project=Project.objects.get(code=project)).order_by('category')
        for sta in stations:
            if sta.script in thisProject_Result.keys():
                if sta.category in stationlist.keys():
                    stationlist[sta.category].append(sta.script)
                else:
                    stationlist[sta.category]=[sta.script]
        
        #查询刷新时间
        new_Git=Git_Object.objects.get(project=project,git=now_git,branch=now_branch)   
        refresh_time=new_Git.pullTime.strftime("%Y-%m-%d %H:%M:%S")

        result_dic['stationlist']=stationlist
        result_dic['branches']=branches
        result_dic['now_branch']=now_branch
        result_dic['gits']=gits
        result_dic['now_git']=now_git
        result_dic['refresh_time']=refresh_time
        
        return result_dic

class PullTC(View): 
    def post(self,request):
        global allStation_Result
        content={}
        branches=[]
        stationlist=[]
        now_branch=''
        refresh_time=""
        res_str=""
        
        project = request.POST['project']
        git=request.POST['git']
        branch=request.POST['branch']
        #查询Git数据表中的地址，来进行pull
        try:
            the_Git=Git_Object.objects.get(project=project,git=git,branch=branch)
            git_adress_psh=os.path.split(os.path.split(the_Git.address)[0])[0]
            
            #git log -1 查询commitId和数据库中作比较
            res=os.popen("cd %s && git pull -f" %git_adress_psh)
            res_log=os.popen("cd %s && git log -1" %git_adress_psh)
            commitId=re.findall('commit\s*(.*?)\s', res_log.read())
        except Exception as e:
            print("pull not sucess!! %s" % e)
            content['result']=False
        else:   

            if (commitId!=[] and commitId[0]==the_Git.latestCommitId):
                the_Git.pullTime=datetime.datetime.now()
                the_Git.save()

                new_Git=Git_Object.objects.get(project=project,git=git,branch=branch)
                content['result']=True
                content['refresh_time']=new_Git.pullTime.strftime("%Y-%m-%d %H:%M:%S")
            else:
                # parse and save data               
                parseTCData(project,git,branch)
                
                #保存新的commit
                the_Git.latestCommitId=commitId[0]
                the_Git.pullTime=datetime.datetime.now()
                the_Git.save()

                new_Git=Git_Object.objects.get(project=project,git=git,branch=branch)
                content['result']=True
                content['refresh_time']=new_Git.pullTime.strftime("%Y-%m-%d %H:%M:%S")
                                     
        return JsonResponse(content, safe=True)


class CreateTCFile(View):
    def post(self,request):      
        allproject_Result=settings.SITE_ALLPROJECTINFO
        project = request.POST['project']
        git=request.POST['git']
        branch=request.POST['branch']
        station = request.POST['station']
        mode = request.POST['mode']
        stage = request.POST['stage']
        File = request.FILES.get("files", None)

        #从文件读取特定分支的数据
        key_str=project+"_"+git+"_"+branch
        thisStation_Result=allproject_Result[key_str]['TC']
        localGitPath=allproject_Result[key_str]['path']

        if File is None:
            return HttpResponse("请选择文件")
        else:
            content = File.read()
            #create test coverage here
            tcContent = self.compentInfoList(thisStation_Result,content.decode(),station,mode)

            mainPath = os.path.join(localGitPath,station,station+'_Main.csv')
            if station == 'TEST2':
                mainPath = os.path.join(localGitPath,'QT0','QT0_Main.csv')
            if station == 'TEST4':
                mainPath = os.path.join(localGitPath,'QT1-PREBURN','QT1-PREBURN_Main.csv')
            mainCotent = ParseGit.getCsvContent(mainPath)

            version= mainCotent[0][1]
            cTime = datetime.datetime.today().strftime("%m%d")
            mode = mode.upper()
            
            downlaodFile = '%s %s_%s_%s_TestCoverage_%s.xlsx' %(station,stage,version,cTime,mode)
            fileName = '%s_%s.xlsx' %(time.time(),station)
            cachePath = '/data/ats/download'
            if not os.path.exists(cachePath):
                os.makedirs(cachePath)
            else:
                fileList=os.listdir(cachePath)
                for f in fileList:
                    os.remove(os.path.join(cachePath,f))
            filePath = os.path.join(cachePath,fileName)
            self.createExcel(filePath,tcContent,station,mainCotent,'Bison')

            file = open(filePath, 'rb')
            response = FileResponse(file)
            response['Content-Type'] = 'application/vnd.ms-excel'
            response['Content-Disposition'] = 'attachment;filename="%s"' % downlaodFile
            return response

    def illegalCharact(self,charact):
        if isinstance(charact,str):
            charact = ILLEGAL_CHARACTERS_RE.sub(r'',charact)
            return charact
        else:
            return charact

    def textToDigit(self,arg):
        if arg.replace('.', '').replace('-', '').isdigit():
            return float(arg)
        else:
            return arg

    def createExcel(self,path, fin_excel_list, name, mainContents, productMark):
        bold_itatic_24_font = Font(name='Arial', size=12, italic=False, color=colors.BLACK, bold=False)
        alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=True, shrink_to_fit=False,
                            indent=0)
        border = Border(left=Side(border_style='thin',
                                color='000000'),
                        right=Side(border_style='thin',
                                color='000000'),
                        top=Side(border_style='thin',
                                color='000000'),
                        bottom=Side(border_style='thin',
                                    color='000000'),
                        diagonal=Side(border_style=None,
                                    color='000000'),
                        diagonal_direction=0,
                        outline=Side(border_style=None,
                                    color='000000'),
                        vertical=Side(border_style=None,
                                    color='000000'),
                        horizontal=Side(border_style=None,
                                        color='000000')
                        )

        if productMark == 'Bison':
            wb = Workbook()
            ws0 = wb.active
            ws1 = wb.create_sheet(name)
            wb.worksheets[0].title = r'Main'

            for item1 in mainContents:
                if len(item1) == 0:
                    continue
                item1[3] = self.textToDigit(item1[3])
                item1[4] = self.textToDigit(item1[4])
                item1[5] = self.textToDigit(item1[5])
                item1[6] = self.textToDigit(item1[6])
                item1[7] = self.textToDigit(item1[7])
                ws0.append(item1)
            endCellIndex = 'H' + str(len(mainContents))

            for row in range(1, len(mainContents) + 1):
                ws0.row_dimensions[row].height = 16
                ws0.column_dimensions['A'].width = 42
                ws0.column_dimensions['B'].width = 24
                ws0.column_dimensions['C'].width = 12
                ws0.column_dimensions['D'].width = 12
                ws0.column_dimensions['E'].width = 12
                ws0.column_dimensions['F'].width = 12
                ws0.column_dimensions['G'].width = 12
                ws0.column_dimensions['H'].width = 12
            cells = ws0['A1':endCellIndex]
            for row_cell in cells:
                for cell in row_cell:
                    cell.border = border
                    cell.font = bold_itatic_24_font
                    cell.alignment = alignment
        else:
            wb = Workbook()
            ws1 = wb.active
            wb.worksheets[0].title = name

        for item in fin_excel_list:
            item[0] = self.illegalCharact(item[0])
            item[1] = self.illegalCharact(item[1])
            item[2] = self.illegalCharact(item[2])
            item[3] = self.textToDigit(item[3])
            item[3] = self.illegalCharact(item[3])
            item[4] = self.textToDigit(item[4])
            item[4] = self.illegalCharact(item[4])
            item[5] = self.textToDigit(item[5])
            item[5] = self.illegalCharact(item[5])
            item[6]=self.illegalCharact(item[6])
            item[7] = self.illegalCharact(item[7])
            ws1.append(item)
        endCellIndex = 'H' + str(len(fin_excel_list))

        for row in range(1, len(fin_excel_list) + 1):
            ws1.row_dimensions[row].height = 16

        ws1.column_dimensions['A'].width = 42
        ws1.column_dimensions['B'].width = 36
        ws1.column_dimensions['C'].width = 10
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 12
        ws1.column_dimensions['G'].width = 36
        ws1.column_dimensions['H'].width = 44

        cells = ws1['A1':endCellIndex]

        for row_cell in cells:
            for cell in row_cell:
                cell.border = border
                cell.font = bold_itatic_24_font
                cell.alignment = alignment

        needFillRows = ws1['A1':'H1']
        fill = PatternFill('solid', fgColor='1E90FF')
        font = Font(name='Arial', size=14, italic=False, color=colors.BLACK, bold=False)
        for row in needFillRows[0]:
            row.fill = fill
            row.font = font
        wb.save(path)

    def catchTimeInLog(self,item,nextItem,source):
        regex = r'(%s,.*?)%s' % (item,nextItem)
        res = re.search(regex,source,re.S)
        if not res:
            return [item,'','','NA','NA','NA','0.00000'] 
        else:
            line = res.group(1)
            lineList=line.split(',')

            return [item,lineList[1],lineList[2],lineList[3],lineList[4],lineList[5],lineList[6]]

    #below for create tc excel
    def compentInfoList(self,thisStation_Result,contentStr,station,mode):
        allItemsList=[['Test Name','Test Commands','Units','lowerLimit','upperLimit','Test Time','Test Actions','Test Parameters']]
        if not contentStr:
            return allItemsList
        stationInfo = thisStation_Result[station]
        index = 0
        while index < len(stationInfo):

            info = stationInfo[index]
            if info[mode] == '1':
                logic = stationInfo[index]['logic']
                name = info['name']
                if index == len(stationInfo) -1:
                    nextName = 'ISN'
                else:
                    nextName = stationInfo[index+1]['name']
                itemList = self.catchTimeInLog(name,nextName,contentStr)

                allItemsList.append([name,logic[0]['TestCommands'],itemList[5],itemList[3],itemList[4],itemList[6],logic[0]['TestActions'],logic[0]['TestParameters']])
                for logicLine in logic[1:]:
                    allItemsList.append(['',logicLine['TestCommands'],'','','','',logicLine['TestActions'],logicLine['TestParameters']])
            index = index + 1
        return allItemsList


def parseTCData(project,git,branch):
    global lock
    the_Git=Git_Object.objects.get(project=project,git=git,branch=branch)
    key_str=project+"_"+git+"_"+branch 
    ParseGit.SearchListByItem = []      #保存搜索要用到的item和command
    ParseGit.SearchListByComm = []      
    settings.SITE_ALLPROJECTINFO[key_str]={"TC":ParseGit.parseLocalGit(the_Git.address),"path":ParseGit.localGitPath,
    "item":ParseGit.SearchListByItem,"command":ParseGit.SearchListByComm}

    #只保存local-Git，master的所有站位信息到数据库中
    if git=='local-git' and branch=='master':
        p = Process(target=SaveData, args=(key_str,project,lock))
        print ('Process will start.')
        p.start()
        # p.join()

#save data into database
def saveProjectData(Station_Info,project,station):
    try:
        try:
            project=Project.objects.get(code=project)
            station=Station.objects.get(script=station,project=Project.objects.get(code=project))
        except Exception as e:
            print(e) 
        else:        
            for i in Station_Info:
                if i['limit']=={}:
                    lower_limit=''
                    upper_limit=''
                    relaxed_lower_limit=''
                    relaxed_upper_limit=''
                    units=''
                    pattern=''
                    materialKey=''
                    materialValue=''
                else:
                    lower_limit = i['limit']['lowerLimit']
                    upper_limit = i['limit']['upperLimit']
                    relaxed_lower_limit = i['limit']['relaxedLowerLimit']
                    relaxed_upper_limit = i['limit']['relaxedUpperLimit']
                    units = i['limit']['units']
                    pattern = i['limit']['pattern']
                    if 'materialKey' in i['limit'].keys():
                        materialKey = i['limit']['materialKey']
                        materialValue = i['limit']['materialValue']
                    else:
                        materialKey = ''
                        materialValue = ''

                stop_on_fail=True if i['stoponfail'] == '1' else False
                disable=True if i['disable'] == '1' else False
                mp = True if i['mp'] == '1' else False
                eng = True if i['eng'] == '1' else False
                rel = True if i['rel'] == '1' else False
                grr = True if i['grr'] == '1' else False
                                
                obj, created = TestCoverage.objects.update_or_create(
                    project=project, station=station, branch='', test_name=i['name'],
                    defaults={"project":project,"station":station,"branch":'',"test_name":i['name'],
                                                            "technology":i['tech'],"stop_on_fail" :stop_on_fail, "disable" : disable,
                                                            "mp" : mp, "eng" : eng, "rel" : rel, "grr" :grr,"lower_limit" : lower_limit,
                                                            "upper_limit" : upper_limit, "relaxed_lower_limit" : relaxed_lower_limit,
                                                            "relaxed_upper_limit" : relaxed_upper_limit, "units" : units, "pattern" : pattern,
                                                            "materialKey" : materialKey, "materialValue" : materialValue, "logic" :i['logic'],
                                                            "description":"test"
                                                            })
    except Exception as e:
        print(e)
        

#创建线程保存所有站位的信息
def SaveData(key_str,project,lock):
    with lock:
        allproject_Result=settings.SITE_ALLPROJECTINFO
        allStation_Result=allproject_Result[key_str]['TC']
        # print(os.getppid())
        # print(os.getpid())
        # print('thread %s ended.' % threading.current_thread().name)

        for sta in allStation_Result.keys():
            thread2=threading.Thread(target=saveProjectData(allStation_Result[sta],project,sta))
            thread2.setDaemon(False)
            thread2.start()


    




